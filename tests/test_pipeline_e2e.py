"""End-to-end pipeline integration test (mock external dependencies).

Validates the back half of the xscreen pipeline (score + report) by feeding
realistic mock Evidence and Ortholog data through the real score.run and
report.run, then asserting on the 4 generated output files.

The front half (search/extract) is mocked away because it depends on the
LLM API and external HTTP services; what we want to verify here is that
the aggregation, scoring, and reporting modules integrate correctly and
produce the expected artifacts.

Score expectations for the mock data (sample_config: min_studies=2,
weight_level=0.5, weight_convergence=0.5, weights T/P/R/F = 1/2/3/4):

    NPF:  4 evidence, 2 studies, levels F+P+T+R -> level_raw=10,
          convergence=2/2=1.0, ortholog present
          total = (0.5*1.0 + 0.5*1.0) * 1.0 = 1.0  <- top
    AKH:  2 evidence, 2 studies, levels F+P      -> level_raw=6,
          convergence=2/2=1.0, ortholog present
          total = (0.5*0.6 + 0.5*1.0) * 1.0 = 0.8
    sNPF:     1 study -> filtered by min_studies=2
    Octopamine: 1 study -> filtered by min_studies=2
"""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src import score, report
from src.extract import Evidence
from src.homolog import Ortholog


# ---------------------------------------------------------------------------
# Fixtures: realistic mock data modeling the locust SIH use case
# ---------------------------------------------------------------------------
@pytest.fixture
def mock_papers():
    """3 mock Paper objects (not used by score/report but document the source)."""
    # Imported here to avoid a hard module-level dependency for the two
    # tests that do not need Paper.
    from src.search import Paper

    return [
        Paper(
            id="P001", source="pdf",
            title="NPF regulates starvation locomotion",
            authors=["Lee"], year=2004, pmid="12345",
            pdf_path="/fake/p001.pdf",
        ),
        Paper(
            id="P002", source="pdf",
            title="AKH and energy homeostasis",
            authors=["Kim"], year=2017, pmid="67890",
            pdf_path="/fake/p002.pdf",
        ),
        Paper(
            id="P003", source="pubmed",
            title="sNPF in Drosophila feeding",
            authors=["Fadda"], year=2019, pmid="11111",
            abstract="sNPF neurons control feeding...",
        ),
    ]


@pytest.fixture
def mock_evidence():
    """8 evidence entries spanning NPF/AKH/sNPF/Octopamine across 3 papers."""
    return [
        # NPF - 2 studies, all four levels (expect top score)
        Evidence(
            id="E1", paper_id="P001", candidate="dNPF", core_name="NPF",
            candidate_type="neuropeptide", species="Drosophila melanogaster",
            evidence_level="functional", direction="down",
            quote="NPF-RNAi increased locomotion under starvation",
            confidence=0.95, source_pmid="12345",
            source_title="NPF regulates starvation locomotion",
            behavior_effect="increased locomotion",
        ),
        Evidence(
            id="E2", paper_id="P001", candidate="dNPF", core_name="NPF",
            candidate_type="neuropeptide", species="Drosophila melanogaster",
            evidence_level="peptide", direction="down",
            quote="NPF immunoreactivity decreased in IPC neurons",
            confidence=0.85, source_pmid="12345",
            source_title="NPF regulates starvation locomotion",
            expression_location="IPC neurons",
        ),
        Evidence(
            id="E3", paper_id="P003", candidate="dNPF", core_name="NPF",
            candidate_type="neuropeptide", species="Drosophila melanogaster",
            evidence_level="transcript", direction="down",
            quote="NPF mRNA reduced after 24h starvation",
            confidence=0.8, source_pmid="11111",
            source_title="sNPF in Drosophila feeding",
        ),
        Evidence(
            id="E4", paper_id="P003", candidate="dNPF", core_name="NPF",
            candidate_type="neuropeptide", species="Drosophila melanogaster",
            evidence_level="release", direction="down",
            quote="NPF release decreased in fed vs starved",
            confidence=0.75, source_pmid="11111",
            source_title="sNPF in Drosophila feeding",
        ),
        # AKH - 2 studies
        Evidence(
            id="E5", paper_id="P002", candidate="AKH", core_name="AKH",
            candidate_type="peptide_hormone", species="Drosophila melanogaster",
            evidence_level="functional", direction="up",
            quote="AKH injection increased locomotion",
            confidence=0.9, source_pmid="67890",
            source_title="AKH and energy homeostasis",
            behavior_effect="hyperactivity",
        ),
        Evidence(
            id="E6", paper_id="P001", candidate="AKH", core_name="AKH",
            candidate_type="peptide_hormone", species="Drosophila melanogaster",
            evidence_level="peptide", direction="up",
            quote="AKH peptide level increased",
            confidence=0.8, source_pmid="12345",
            source_title="NPF regulates starvation locomotion",
        ),
        # sNPF - 1 study (filtered by min_studies=2)
        Evidence(
            id="E7", paper_id="P003", candidate="sNPF", core_name="sNPF",
            candidate_type="neuropeptide", species="Drosophila melanogaster",
            evidence_level="transcript", direction="up",
            quote="sNPF expression increased",
            confidence=0.7, source_pmid="11111",
            source_title="sNPF in Drosophila feeding",
        ),
        # Octopamine - 1 study (filtered)
        Evidence(
            id="E8", paper_id="P002", candidate="Octopamine",
            core_name="Octopamine",
            candidate_type="biogenic_amine",
            species="Drosophila melanogaster",
            evidence_level="functional", direction="up",
            quote="Octopamine necessary for starvation hyperactivity",
            confidence=0.95, source_pmid="67890",
            source_title="AKH and energy homeostasis",
            behavior_effect="hyperactivity",
        ),
    ]


@pytest.fixture
def mock_orthologs():
    """Ortholog map: NPF/AKH/Octopamine mapped, sNPF has no ortholog."""
    return {
        "NPF": Ortholog(
            source_gene="NPF", target_gene="NPF1a",
            identity=0.7, coverage=0.7,
            source_species="Drosophila melanogaster",
            target_species="Locusta migratoria",
            uniprot_id="A0A0",
        ),
        "AKH": Ortholog(
            source_gene="AKH", target_gene="AKH-like",
            identity=0.7, coverage=0.7,
            source_species="Drosophila melanogaster",
            target_species="Locusta migratoria",
            uniprot_id="B0B0",
        ),
        "sNPF": None,
        "Octopamine": Ortholog(
            source_gene="Octopamine", target_gene="TyrR",
            identity=0.7, coverage=0.7,
            source_species="Drosophila melanogaster",
            target_species="Locusta migratoria",
            uniprot_id="C0C0",
        ),
    }


# ---------------------------------------------------------------------------
# 1. All four output files produced
# ---------------------------------------------------------------------------
def test_pipeline_produces_all_outputs(
    sample_config, mock_evidence, mock_orthologs, tmp_path
):
    sample_config["output"]["dir"] = str(tmp_path)
    scores = score.run(sample_config, mock_evidence, mock_orthologs)
    report.run(sample_config, Path(tmp_path), scores, mock_evidence)

    for fname in (
        "candidates_ranked.xlsx",
        "evidence_detail.xlsx",
        "evidence_db.json",
        "report.md",
    ):
        assert (tmp_path / fname).exists(), f"missing output: {fname}"


# ---------------------------------------------------------------------------
# 2. NPF ranks #1 (4 evidence, 2 studies, all levels, ortholog present)
# ---------------------------------------------------------------------------
def test_pipeline_npf_ranks_top(
    sample_config, mock_evidence, mock_orthologs
):
    scores = score.run(sample_config, mock_evidence, mock_orthologs)
    assert len(scores) >= 1
    assert scores[0].candidate == "NPF"
    # Sanity-check the exact score value: (0.5*1.0 + 0.5*1.0) * 1.0 = 1.0
    assert abs(scores[0].total_score - 1.0) < 1e-9


# ---------------------------------------------------------------------------
# 3. min_studies filter removes single-study candidates
# ---------------------------------------------------------------------------
def test_pipeline_min_studies_filter(
    sample_config, mock_evidence, mock_orthologs
):
    scores = score.run(sample_config, mock_evidence, mock_orthologs)
    candidates = {s.candidate for s in scores}
    # NPF (2 studies) and AKH (2 studies) survive; sNPF and Octopamine
    # each have only 1 study and must be filtered out.
    assert "NPF" in candidates
    assert "AKH" in candidates
    assert "sNPF" not in candidates
    assert "Octopamine" not in candidates
    for s in scores:
        assert s.study_count >= sample_config["scoring"]["min_studies"]


# ---------------------------------------------------------------------------
# 4. Excel ranked table: NPF on the first data row
# ---------------------------------------------------------------------------
def test_pipeline_excel_has_ranked_data(
    sample_config, mock_evidence, mock_orthologs, tmp_path
):
    sample_config["output"]["dir"] = str(tmp_path)
    scores = score.run(sample_config, mock_evidence, mock_orthologs)
    report.run(sample_config, Path(tmp_path), scores, mock_evidence)

    wb = load_workbook(tmp_path / "candidates_ranked.xlsx")
    ws = wb.active
    # Header row
    headers = [c.value for c in ws[1]]
    assert "Rank" in headers
    assert "Candidate" in headers

    # First data row: rank 1, candidate name starting with dNPF (original name)
    rank_cell = ws.cell(row=2, column=headers.index("Rank") + 1).value
    candidate_cell = ws.cell(row=2, column=headers.index("Candidate") + 1).value
    core_cell = ws.cell(row=2, column=headers.index("Core name") + 1).value
    assert rank_cell == 1
    assert core_cell == "NPF"
    assert candidate_cell == "dNPF"

    # Exactly two candidates (NPF + AKH) survived the min_studies filter.
    data_rows = ws.max_row - 1
    assert data_rows == 2


# ---------------------------------------------------------------------------
# 5. Markdown report mentions the ranked candidates
# ---------------------------------------------------------------------------
def test_pipeline_md_has_candidates(
    sample_config, mock_evidence, mock_orthologs, tmp_path
):
    sample_config["output"]["dir"] = str(tmp_path)
    scores = score.run(sample_config, mock_evidence, mock_orthologs)
    report.run(sample_config, Path(tmp_path), scores, mock_evidence)

    md = (tmp_path / "report.md").read_text(encoding="utf-8")
    assert "NPF" in md
    assert "AKH" in md
    # Methodology section should document the filters used
    assert "min_studies" in md
    # Exactly two ranked candidates (NPF + AKH). Filtered candidates
    # (sNPF, Octopamine) must not appear as ranked rows. The substring
    # "sNPF" can legitimately appear inside a source_title quote
    # ("sNPF in Drosophila feeding"), so we check the ranked-table rows
    # explicitly instead of the whole document.
    table_rows = [
        line for line in md.splitlines()
        if line.startswith("| ") and "Rank" not in line and "---" not in line
    ]
    ranked_core_names = {row.split("|")[2].strip() for row in table_rows}
    assert "NPF" in ranked_core_names
    assert "AKH" in ranked_core_names
    assert "sNPF" not in ranked_core_names
    assert "Octopamine" not in ranked_core_names


# ---------------------------------------------------------------------------
# 6. JSON audit trail: non-empty candidates + evidence arrays
# ---------------------------------------------------------------------------
def test_pipeline_json_audit_trail(
    sample_config, mock_evidence, mock_orthologs, tmp_path
):
    sample_config["output"]["dir"] = str(tmp_path)
    scores = score.run(sample_config, mock_evidence, mock_orthologs)
    report.run(sample_config, Path(tmp_path), scores, mock_evidence)

    data = json.loads((tmp_path / "evidence_db.json").read_text())
    assert "candidates" in data
    assert "evidence" in data
    # Two candidates survived scoring
    assert len(data["candidates"]) == 2
    # All 8 input evidence entries are preserved for traceability
    assert len(data["evidence"]) == 8
    candidate_names = {c["candidate"] for c in data["candidates"]}
    assert candidate_names == {"NPF", "AKH"}
    # NPF should be first in the list (rank 1)
    assert data["candidates"][0]["candidate"] == "NPF"


# ---------------------------------------------------------------------------
# 7. Full src.run.main pipeline, fully mocked (search/extract/homolog)
# ---------------------------------------------------------------------------
def test_pipeline_full_main_mocked(
    sample_config, mock_papers, mock_evidence, mock_orthologs, tmp_path,
    monkeypatch,
):
    """Drive src.run.main end-to-end with every external dependency mocked.

    Mocks:
      - config_loader.load_config -> sample_config with tmp output dir
      - config_loader.get_output_dir -> tmp_path
      - search.run -> mock_papers
      - extract.run -> mock_evidence
      - homolog.run -> mock_orthologs
    Verifies all 4 output files exist after the full orchestrator runs.
    """
    from src import config_loader, run as run_module

    sample_config["output"]["dir"] = str(tmp_path)

    monkeypatch.setattr(
        config_loader, "load_config", lambda path: sample_config
    )
    monkeypatch.setattr(
        config_loader, "get_output_dir",
        lambda config, path: Path(tmp_path),
    )
    monkeypatch.setattr(
        "src.run.search.run", lambda config: mock_papers, raising=True
    )
    monkeypatch.setattr(
        "src.run.extract.run",
        lambda config, papers: mock_evidence,
        raising=True,
    )
    monkeypatch.setattr(
        "src.run.homolog.run",
        lambda config, candidates: mock_orthologs,
        raising=True,
    )

    run_module.main("/fake/config.yaml")

    for fname in (
        "candidates_ranked.xlsx",
        "evidence_detail.xlsx",
        "evidence_db.json",
        "report.md",
    ):
        assert (tmp_path / fname).exists(), f"main() missing output: {fname}"
