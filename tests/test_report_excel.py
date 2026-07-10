"""Tests for Excel report generation (src/report.py)."""
import openpyxl

from src.extract import Evidence
from src.homolog import Ortholog
from src.report import write_excel, write_evidence_detail, run
from src.score import CandidateScore


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _make_score(cand: str = "NPF", ortholog: Ortholog | None = None) -> CandidateScore:
    return CandidateScore(
        candidate=cand,
        total_score=0.8,
        evidence_count=3,
        study_count=2,
        evidence_levels={"transcript": 1, "functional": 2},
        direction_consistency=0.9,
        ortholog=ortholog,
        top_papers=["PM1", "PM2"],
        score_breakdown={
            "level_raw": 9,
            "level_norm": 1.0,
            "convergence": 0.67,
            "ortholog_mult": 1.0,
            "w_level": 0.5,
            "w_conv": 0.5,
        },
    )


def _make_ortholog(gene: str = "LmNPF") -> Ortholog:
    return Ortholog(
        source_gene="NPF",
        target_gene=gene,
        identity=0.7,
        coverage=0.7,
        source_species="Drosophila melanogaster",
        target_species="Locusta migratoria",
        uniprot_id="Q12345",
    )


def _make_evidence(
    eid: str = "E1",
    core_name: str = "NPF",
    candidate: str = "dNPF",
    candidate_type: str = "neuropeptide",
    pmid: str = "PM1",
    level: str = "transcript",
    direction: str = "up",
) -> Evidence:
    return Evidence(
        id=eid,
        paper_id="P1",
        candidate=candidate,
        core_name=core_name,
        candidate_type=candidate_type,
        species="Drosophila melanogaster",
        evidence_level=level,
        direction=direction,
        quote="NPF expression increased",
        confidence=0.9,
        source_pmid=pmid,
        source_title="Test paper",
        behavior_effect="increased locomotion",
        expression_location="IPC neurons",
    )


# ---------------------------------------------------------------------------
# 1. Main table columns and first-row data
# ---------------------------------------------------------------------------
def test_write_excel_columns(tmp_path):
    """write_excel produces 11 columns with correct headers and first-row data."""
    ortho = _make_ortholog()
    score = _make_score(cand="NPF", ortholog=ortho)
    ev = _make_evidence(core_name="NPF", candidate="dNPF")
    out = tmp_path / "candidates_ranked.xlsx"

    write_excel([score], [ev], out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == [
        "Rank", "Candidate", "Core name", "Type", "Ortholog (target)",
        "Total score", "Level score", "Convergence", "Studies",
        "Evidence levels", "Key refs",
    ]
    # First data row
    row = [c.value for c in ws[2]]
    assert row[0] == 1            # Rank
    assert row[1] == "dNPF"       # Candidate (original name from evidence)
    assert row[2] == "NPF"        # Core name
    assert row[3] == "neuropeptide"  # Type
    assert row[4] == "LmNPF"      # Ortholog target
    assert row[5] == 0.8          # Total score
    assert row[6] == 9            # Level score (level_raw)
    assert row[7] == 0.67         # Convergence
    assert row[8] == 2            # Studies
    assert "T:1" in row[9]        # Evidence levels contains T:1
    assert "F:2" in row[9]
    assert row[10] == "PM1, PM2"  # Key refs


# ---------------------------------------------------------------------------
# 2. No ortholog => "not found"
# ---------------------------------------------------------------------------
def test_write_excel_no_ortholog(tmp_path):
    score = _make_score(cand="NPF", ortholog=None)
    ev = _make_evidence(core_name="NPF")
    out = tmp_path / "out.xlsx"

    write_excel([score], [ev], out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    row = [c.value for c in ws[2]]
    assert row[4] == "not found"  # Ortholog column


# ---------------------------------------------------------------------------
# 3. Evidence levels format
# ---------------------------------------------------------------------------
def test_write_excel_evidence_levels_format(tmp_path):
    score = CandidateScore(
        candidate="NPF",
        total_score=0.5,
        evidence_count=3,
        study_count=2,
        evidence_levels={"transcript": 1, "functional": 2},
        direction_consistency=1.0,
        ortholog=_make_ortholog(),
        top_papers=["PM1"],
        score_breakdown={
            "level_raw": 9, "level_norm": 1.0, "convergence": 0.5,
            "ortholog_mult": 1.0, "w_level": 0.5, "w_conv": 0.5,
        },
    )
    ev = _make_evidence(core_name="NPF")
    out = tmp_path / "out.xlsx"

    write_excel([score], [ev], out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    levels_cell = ws.cell(row=2, column=10).value
    parts = levels_cell.split()
    assert "T:1" in parts
    assert "F:2" in parts
    # No zero-count levels
    for p in parts:
        assert not p.endswith(":0")


# ---------------------------------------------------------------------------
# 4. Evidence detail table
# ---------------------------------------------------------------------------
def test_write_evidence_detail_columns(tmp_path):
    ev1 = _make_evidence(eid="E1", core_name="NPF", candidate="dNPF", pmid="PM1")
    ev2 = _make_evidence(
        eid="E2", core_name="NPF", candidate="NPF", pmid="PM2",
        level="functional", direction="down",
    )
    out = tmp_path / "evidence_detail.xlsx"

    write_evidence_detail([ev1, ev2], out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == [
        "Core name", "Candidate", "Paper ID", "PMID", "Level",
        "Direction", "Behavior", "Location", "Quote", "Confidence",
    ]
    assert ws.max_row == 3  # header + 2 data rows
    row1 = [c.value for c in ws[2]]
    assert row1[0] == "NPF"
    assert row1[1] == "dNPF"
    assert row1[2] == "P1"
    assert row1[3] == "PM1"
    assert row1[4] == "transcript"
    assert row1[5] == "up"
    assert row1[6] == "increased locomotion"
    assert row1[7] == "IPC neurons"
    assert row1[8] == "NPF expression increased"
    assert row1[9] == 0.9


# ---------------------------------------------------------------------------
# 5. report.run creates all three files
# ---------------------------------------------------------------------------
def test_report_run_creates_files(tmp_path, sample_config):
    """run() should produce candidates_ranked.xlsx, evidence_detail.xlsx,
    evidence_db.json, and report.md (all four output formats)."""
    cfg = {**sample_config}
    cfg["output"] = {**sample_config["output"], "evidence_detail": "evidence_detail.xlsx"}

    ortho = _make_ortholog()
    score = _make_score(cand="NPF", ortholog=ortho)
    ev = _make_evidence(core_name="NPF")

    run(cfg, tmp_path, [score], [ev])

    assert (tmp_path / "candidates_ranked.xlsx").exists()
    assert (tmp_path / "evidence_detail.xlsx").exists()
    assert (tmp_path / "evidence_db.json").exists()
    assert (tmp_path / "report.md").exists()


# ---------------------------------------------------------------------------
# 6. Type inference from multiple evidence entries
# ---------------------------------------------------------------------------
def test_type_inference(tmp_path):
    """Same candidate with multiple evidence entries of same type; Type column
    reflects that type."""
    score = _make_score(cand="AKH")
    evs = [
        _make_evidence(eid="E1", core_name="AKH", candidate="dAKH", candidate_type="neuropeptide"),
        _make_evidence(eid="E2", core_name="AKH", candidate="AKH", candidate_type="neuropeptide"),
    ]
    out = tmp_path / "out.xlsx"

    write_excel([score], evs, out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    type_cell = ws.cell(row=2, column=4).value
    assert type_cell == "neuropeptide"
